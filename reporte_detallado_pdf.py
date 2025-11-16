import os
from datetime import datetime, date
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ReporteDetalladoPDF:
    def __init__(self, db):
        self.db = db

    # -------------------------
    # Utilidades internas
    # -------------------------

    @staticmethod
    def _normalizar_fecha_para_str(valor, default_label):
        """
        Normaliza una fecha (puede venir como datetime, date, str o None)
        y devuelve siempre un string legible.

        - Si es datetime/date: se formatea con '%Y-%m-%d'.
        - Si es str: se devuelve tal cual (asumiendo que ya viene formateada).
        - Si es None o vacío: devuelve default_label (p.ej. 'Inicio' o 'Fin').
        """
        if not valor:
            return default_label

        if isinstance(valor, (datetime, date)):
            return valor.strftime("%Y-%m-%d")

        # Si ya es string, lo usamos tal cual
        if isinstance(valor, str):
            return valor

        # Caso raro: otro tipo -> lo convertimos a string
        return str(valor)

    @staticmethod
    def _extraer_fechas_de_filtros(filtros):
        """
        Extrae fecha_inicio y fecha_fin desde el dict de filtros y devuelve
        dos strings ya normalizados para usar en títulos/encabezados.
        """
        fecha_ini_val = filtros.get("fecha_inicio")
        fecha_fin_val = filtros.get("fecha_fin")

        fecha_ini_str = ReporteDetalladoPDF._normalizar_fecha_para_str(
            fecha_ini_val, "Inicio"
        )
        fecha_fin_str = ReporteDetalladoPDF._normalizar_fecha_para_str(
            fecha_fin_val, "Fin"
        )
        return fecha_ini_str, fecha_fin_str

    # -------------------------
    # Exportar a PDF
    # -------------------------

    def exportar(
        self,
        proyecto_id,
        filtros,
        cliente_nombre,
        moneda_symbol="RD$",
        nombre_archivo=None,
        ruta_forzada=None,
    ):
        """
        Genera el PDF detallado de alquileres.
        - proyecto_id: ID del proyecto.
        - filtros: dict de filtros (fecha_inicio, fecha_fin, cliente_id, etc).
        - cliente_nombre: nombre del cliente seleccionado, o 'TODOS LOS CLIENTES'.
        - moneda_symbol: símbolo de la moneda (ej: 'RD$')
        - nombre_archivo: opcional, si se quiere un nombre específico para el PDF
        - ruta_forzada: si se quiere especificar la ruta absoluta del archivo.
        """
        # 1. Selección de ruta de guardado
        if ruta_forzada:
            ruta_guardar = ruta_forzada
        else:
            if not nombre_archivo:
                nombre_archivo = (
                    f"Reporte_Detallado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                )
            ruta_guardar = nombre_archivo

        # 2. Obtención de datos
        datos = self.db.obtener_transacciones_por_proyecto(proyecto_id, filtros)
        if not datos:
            return False, "No hay datos para generar el reporte PDF."

        # 3. Construcción del PDF
        doc = SimpleDocTemplate(ruta_guardar, pagesize=letter)
        estilos = getSampleStyleSheet()
        estilos.add(ParagraphStyle(name="RightAlign", alignment=TA_RIGHT))
        elementos = []

        # Fechas de filtro normalizadas (evita error de strftime sobre str)
        fecha_ini_str, fecha_fin_str = self._extraer_fechas_de_filtros(filtros)

        elementos.append(
            Paragraph("REPORTE DE ALQUILER EQUIPOS PESADOS", estilos["h1"])
        )
        elementos.append(Paragraph(f"Cliente: {cliente_nombre}", estilos["Normal"]))
        elementos.append(
            Paragraph(f"Periodo: {fecha_ini_str} al {fecha_fin_str}", estilos["Normal"])
        )
        elementos.append(
            Paragraph(
                f"Fecha de generación: "
                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                estilos["Normal"],
            )
        )
        elementos.append(Spacer(1, 0.25 * inch))

        # 4. Procesamiento y tablas por equipo
        df = pd.DataFrame([dict(row) for row in datos])

        # Seguridad: columnas que vamos a usar
        for col in ["equipo_nombre", "horas", "monto", "cliente_nombre", "ubicacion"]:
            if col not in df.columns:
                df[col] = None

        df["equipo_nombre"] = df["equipo_nombre"].fillna("Sin Equipo")
        equipos = df["equipo_nombre"].unique()
        totales_resumen = []

        for eq in equipos:
            elementos.append(Paragraph(f"Equipo: {eq.upper()}", estilos["h2"]))

            df_equipo = df[df["equipo_nombre"] == eq].copy()

            # Aseguramos que horas y monto sean numéricos para evitar errores al sumar
            df_equipo["horas"] = pd.to_numeric(df_equipo["horas"], errors="coerce").fillna(0)
            df_equipo["monto"] = pd.to_numeric(df_equipo["monto"], errors="coerce").fillna(0)

            tabla_data = [["Fecha", "Conduce", "Horas", "Cliente/Ubicación", "Monto"]]
            total_horas_eq = df_equipo["horas"].sum()
            total_monto_eq = df_equipo["monto"].sum()

            for _, row in df_equipo.iterrows():
                cliente_nombre_val = row.get("cliente_nombre") or ""
                ubicacion_val = row.get("ubicacion") or ""
                if cliente_nombre_val and ubicacion_val:
                    ubicacion_cliente = f"{cliente_nombre_val} / {ubicacion_val}"
                elif cliente_nombre_val:
                    ubicacion_cliente = cliente_nombre_val
                else:
                    ubicacion_cliente = ubicacion_val or "N/A"

                fecha_str = str(row.get("fecha", ""))
                conduce_str = str(row.get("conduce", ""))

                horas_val = float(row.get("horas", 0) or 0)
                monto_val = float(row.get("monto", 0) or 0)

                tabla_data.append(
                    [
                        fecha_str,
                        conduce_str,
                        f"{horas_val:.2f}",
                        Paragraph(ubicacion_cliente, estilos["Normal"]),
                        f"{moneda_symbol} {monto_val:,.2f}",
                    ]
                )

            # Totales por equipo
            tabla_data.append(
                [
                    "",
                    "",
                    "",
                    Paragraph("<b>Total Horas:</b>", estilos["Normal"]),
                    Paragraph(f"<b>{total_horas_eq:.2f}</b>", estilos["Normal"]),
                ]
            )
            tabla_data.append(
                [
                    "",
                    "",
                    "",
                    Paragraph("<b>Total Monto:</b>", estilos["Normal"]),
                    Paragraph(
                        f"<b>{moneda_symbol} {total_monto_eq:,.2f}</b>",
                        estilos["Normal"],
                    ),
                ]
            )

            totales_resumen.append([eq, total_horas_eq, total_monto_eq])

            tabla = Table(
                tabla_data,
                colWidths=[0.8 * inch, 0.8 * inch, 0.7 * inch, 3 * inch, 1.2 * inch],
            )
            tabla.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6321")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ALIGN", (3, 1), (3, -1), "LEFT"),
                        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                        ("ALIGN", (3, -2), (3, -1), "RIGHT"),
                    ]
                )
            )
            elementos.append(tabla)
            elementos.append(Spacer(1, 0.2 * inch))

        # 5. Resumen general por equipos
        elementos.append(Paragraph("RESUMEN GENERAL POR EQUIPOS", estilos["h2"]))
        resumen_data = [["Equipo", "Total Horas", "Total Monto"]]

        total_general_horas = 0.0
        total_general_monto = 0.0
        for eq, horas, monto in sorted(totales_resumen, key=lambda x: x[0]):
            horas_val = float(horas or 0)
            monto_val = float(monto or 0)
            resumen_data.append(
                [
                    Paragraph(eq, estilos["Normal"]),
                    f"{horas_val:.2f}",
                    f"{moneda_symbol} {monto_val:,.2f}",
                ]
            )
            total_general_horas += horas_val
            total_general_monto += monto_val

        resumen_data.append(
            [
                Paragraph("<b>TOTAL GENERAL</b>", estilos["Normal"]),
                Paragraph(f"<b>{total_general_horas:.2f}</b>", estilos["RightAlign"]),
                Paragraph(
                    f"<b>{moneda_symbol} {total_general_monto:,.2f}</b>",
                    estilos["RightAlign"],
                ),
            ]
        )

        tabla_resumen = Table(
            resumen_data, colWidths=[3 * inch, 1.5 * inch, 2 * inch]
        )
        tabla_resumen.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6321")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("BACKGROUND", (-3, -1), (-1, -1), colors.lightgrey),
                ]
            )
        )
        elementos.append(tabla_resumen)

        # 6. Guardar el PDF
        try:
            doc.build(elementos)
            return True, ruta_guardar
        except PermissionError:
            return (
                False,
                "No se pudo guardar el archivo. Asegúrate de que no esté abierto.",
            )
        except Exception as e:
            return False, str(e)

    # -------------------------
    # Exportar a Excel
    # -------------------------

    def exportar_excel(
        self,
        proyecto_id,
        filtros,
        cliente_nombre,
        moneda_symbol="RD$",
        nombre_archivo=None,
        ruta_forzada=None,
    ):
        """
        Exporta el reporte detallado de alquileres a un archivo Excel,
        con formato profesional y totales.
        Permite nombre personalizado y ruta forzada como el PDF.
        """
        # 1. Selección de ruta de guardado
        if ruta_forzada:
            ruta_guardar = ruta_forzada
        else:
            if not nombre_archivo:
                nombre_archivo = (
                    f"Reporte_Detallado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
            ruta_guardar = nombre_archivo

        # 2. Obtención de datos
        datos = self.db.obtener_transacciones_por_proyecto(proyecto_id, filtros)
        if not datos:
            return (
                False,
                "No hay datos que coincidan con los filtros para generar el reporte.",
            )

        df = pd.DataFrame([dict(row) for row in datos])

        # 3. Excel: encabezados, formato y datos
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Detallado"

        # Título
        ws.merge_cells("A1:J1")
        ws["A1"] = "Reporte Detallado de Alquiler de Equipos"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(
            start_color="002060", end_color="002060", fill_type="solid"
        )
        ws.row_dimensions[1].height = 30

        # Período
        ws.merge_cells("A2:J2")
        fecha_ini_str, fecha_fin_str = self._extraer_fechas_de_filtros(filtros)
        ws["A2"] = f"Período del reporte: {fecha_ini_str} al {fecha_fin_str}"
        ws["A2"].font = Font(size=11, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # Columnas a exportar
        columnas_export = [
            "fecha",
            "conduce",
            "cliente_nombre",
            "operador_nombre",
            "equipo_nombre",
            "ubicacion",
            "horas",
            "precio_por_hora",
            "monto",
            "pagado",
        ]
        columnas_renombradas = [
            "Fecha",
            "Conduce",
            "Cliente",
            "Operador",
            "Equipo",
            "Ubicación",
            "Horas",
            "Precio/Hora",
            "Monto",
            "Estado",
        ]

        # Aseguramos la existencia de columnas
        for col in columnas_export:
            if col not in df.columns:
                df[col] = None

        df_export = df[columnas_export].copy()
        df_export["Estado"] = df_export["pagado"].apply(
            lambda x: "Pagado" if x else "Pendiente"
        )
        df_export = df_export.drop(columns=["pagado"])
        df_export.columns = columnas_renombradas

        # Escribir DataFrame a la hoja
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws.append(r)

        # Estilo de encabezados (fila 3)
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Formatos numéricos
        formato_horas = "0.00"
        formato_moneda = f'"{moneda_symbol}" #,##0.00'

        # Asumimos que las columnas G, H, I corresponden a Horas, Precio/Hora, Monto
        for row_idx in range(4, ws.max_row + 1):
            ws[f"G{row_idx}"].number_format = formato_horas
            ws[f"H{row_idx}"].number_format = formato_moneda
            ws[f"I{row_idx}"].number_format = formato_moneda

        # Cálculo de totales
        df["monto"] = pd.to_numeric(df["monto"], errors="coerce").fillna(0)
        df["horas"] = pd.to_numeric(df["horas"], errors="coerce").fillna(0)

        total_facturado = df["monto"].sum()
        total_abonado = df[df["pagado"] == 1]["monto"].sum()
        total_pendiente = total_facturado - total_abonado
        total_horas = df["horas"].sum()

        fila_inicio_totales = ws.max_row + 2

        ws.cell(row=fila_inicio_totales, column=8, value="Total Facturado:").font = Font(
            bold=True
        )
        ws.cell(
            row=fila_inicio_totales + 1,
            column=8,
            value="Total Pagado:",
        ).font = Font(bold=True, color="00B050")
        ws.cell(
            row=fila_inicio_totales + 2,
            column=8,
            value="Total Pendiente:",
        ).font = Font(bold=True, color="FF0000")
        ws.cell(
            row=fila_inicio_totales + 3,
            column=8,
            value="Total Horas:",
        ).font = Font(bold=True)

        ws.cell(
            row=fila_inicio_totales,
            column=9,
            value=total_facturado,
        ).number_format = formato_moneda
        ws.cell(
            row=fila_inicio_totales + 1,
            column=9,
            value=total_abonado,
        ).number_format = formato_moneda
        ws.cell(
            row=fila_inicio_totales + 2,
            column=9,
            value=total_pendiente,
        ).number_format = formato_moneda
        ws.cell(
            row=fila_inicio_totales + 3,
            column=9,
            value=total_horas,
        ).number_format = formato_horas

        # Ajuste de ancho de columnas
        for col in ws.iter_cols(min_row=3):
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar Excel
        try:
            wb.save(ruta_guardar)
            return True, ruta_guardar
        except PermissionError:
            return (
                False,
                "No se pudo guardar el archivo. Asegúrate de que no esté abierto.",
            )
        except Exception as e:
            return False, str(e)