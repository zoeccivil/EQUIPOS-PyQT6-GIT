from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

class ReporteOperadores:
    def __init__(self, db):
        self.db = db

    def exportar_pdf(self, proyecto_id, filtros, ruta_guardar, moneda_symbol='RD$'):
        # --- 1. Obtener datos ---
        ingresos_data = self.db.obtener_transacciones_por_proyecto(proyecto_id, filtros)
        pagos_data = self.db.obtener_pagos_a_operadores(proyecto_id, filtros)
        if not ingresos_data:
            return False, "No hay datos de ingresos para generar el reporte PDF."

        doc = SimpleDocTemplate(ruta_guardar, pagesize=landscape(letter))
        estilos = getSampleStyleSheet()
        estilos.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT))
        estilos.add(ParagraphStyle(name='RightAlignBold', fontName='Helvetica-Bold', alignment=TA_RIGHT))
        estilos.add(ParagraphStyle(name='CenterBold', fontName='Helvetica-Bold', alignment=TA_CENTER))
        estilos.add(ParagraphStyle(name='TblRight', fontSize=9, alignment=TA_RIGHT))
        estilos.add(ParagraphStyle(name='TblLeft', fontSize=9, alignment=TA_LEFT))
        estilos.add(ParagraphStyle(name='TblCenter', fontSize=9, alignment=TA_CENTER))

        elementos = []
        fecha_ini_str = filtros.get('fecha_inicio', datetime.now()).strftime('%Y-%m-%d') if filtros.get('fecha_inicio') else "Inicio"
        fecha_fin_str = filtros.get('fecha_fin', datetime.now()).strftime('%Y-%m-%d') if filtros.get('fecha_fin') else "Fin"
        elementos.append(Paragraph(f"Resumen Financiero de Operadores", estilos['h1']))
        elementos.append(Paragraph(f"Período: {fecha_ini_str} al {fecha_fin_str}", estilos['Normal']))
        elementos.append(Spacer(1, 0.2 * inch))

        # 2. Procesar con pandas
        df = pd.DataFrame([dict(row) for row in ingresos_data])
        if df.empty:
            return False, "No hay datos para generar el reporte PDF."
        df['operador_nombre'] = df['operador_nombre'].fillna('No Especificado')
        df['equipo_nombre'] = df['equipo_nombre'].fillna('No Especificado')

        # Resumen por operador y equipo
        resumen_ingresos = df.groupby(['operador_nombre', 'equipo_nombre']).agg(
            total_horas=('horas', 'sum'),
            total_ingresos=('monto', 'sum')
        ).reset_index()

        # Pagos
        if pagos_data:
            df_pagos = pd.DataFrame([dict(row) for row in pagos_data])
            resumen_pagos = df_pagos.groupby(['operador_nombre', 'equipo_nombre']).agg(
                total_pagado=('monto', 'sum')
            ).reset_index()
            resumen_final = pd.merge(resumen_ingresos, resumen_pagos, on=['operador_nombre', 'equipo_nombre'], how='left')
        else:
            resumen_final = resumen_ingresos
            resumen_final['total_pagado'] = 0.0

        resumen_final['total_pagado'] = resumen_final['total_pagado'].fillna(0)
        resumen_final['tarifa_efectiva'] = resumen_final.apply(
            lambda row: (row['total_pagado'] / row['total_horas']) if row['total_horas'] > 0 else 0, axis=1
        )

        # Tabla PDF
        tabla_data = [[
            Paragraph("<b>Operador</b>", estilos['TblCenter']),
            Paragraph("<b>Equipo</b>", estilos['TblCenter']),
            Paragraph("<b>Total Horas</b>", estilos['TblCenter']),
            Paragraph("<b>Ingreso Generado</b>", estilos['TblCenter']),
            Paragraph("<b>Total Pagado</b>", estilos['TblCenter']),
            Paragraph(f"<b>Tarifa Efectiva ({moneda_symbol}/hr)</b>", estilos['TblCenter'])
        ]]
        for _, row in resumen_final.sort_values(by=['operador_nombre', 'equipo_nombre']).iterrows():
            tabla_data.append([
                Paragraph(str(row['operador_nombre']), estilos['TblLeft']),
                Paragraph(str(row['equipo_nombre']), estilos['TblLeft']),
                Paragraph(f"{row['total_horas']:.2f}", estilos['TblRight']),
                Paragraph(f"{moneda_symbol} {row['total_ingresos']:,.2f}", estilos['TblRight']),
                Paragraph(f"{moneda_symbol} {row['total_pagado']:,.2f}", estilos['TblRight']),
                Paragraph(f"{moneda_symbol} {row['tarifa_efectiva']:,.2f}", estilos['TblRight']),
            ])

        # Fila de Totales
        tabla_data.append([
            Paragraph("<b>TOTALES</b>", estilos['TblLeft']),
            "",
            Paragraph(f"<b>{resumen_final['total_horas'].sum():.2f}</b>", estilos['RightAlignBold']),
            Paragraph(f"<b>{moneda_symbol} {resumen_final['total_ingresos'].sum():,.2f}</b>", estilos['RightAlignBold']),
            Paragraph(f"<b>{moneda_symbol} {resumen_final['total_pagado'].sum():,.2f}</b>", estilos['RightAlignBold']),
            ""
        ])

        tabla = Table(tabla_data, colWidths=[2.5*inch,2*inch,1.2*inch,1.6*inch,1.6*inch,1.5*inch], repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#003366")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (0,1), (1,-1), 'LEFT'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ]))
        elementos.append(tabla)

        try:
            doc.build(elementos)
            return True, ruta_guardar
        except PermissionError:
            return False, "No se pudo guardar el archivo. Asegúrate de que no esté abierto."
        except Exception as e:
            return False, str(e)

    def exportar_excel(self, proyecto_id, filtros, ruta_guardar, moneda_symbol="RD$"):
        ingresos_data = self.db.obtener_transacciones_por_proyecto(proyecto_id, filtros)
        pagos_data = self.db.obtener_pagos_a_operadores(proyecto_id, filtros)
        if not ingresos_data:
            return False, "No hay datos que coincidan con los filtros para generar el reporte."

        df = pd.DataFrame([dict(row) for row in ingresos_data])
        df['operador_nombre'] = df['operador_nombre'].fillna('No Especificado')
        df['equipo_nombre'] = df['equipo_nombre'].fillna('No Especificado')
        resumen_ingresos = df.groupby(['operador_nombre','equipo_nombre']).agg(
            total_horas=('horas','sum'),
            total_ingresos=('monto','sum')
        ).reset_index()
        if pagos_data:
            df_pagos = pd.DataFrame([dict(row) for row in pagos_data])
            resumen_pagos = df_pagos.groupby(['operador_nombre','equipo_nombre']).agg(
                total_pagado=('monto','sum')
            ).reset_index()
            resumen_final = pd.merge(resumen_ingresos, resumen_pagos, on=['operador_nombre','equipo_nombre'], how='left')
        else:
            resumen_final = resumen_ingresos
            resumen_final['total_pagado'] = 0.0

        resumen_final['total_pagado'] = resumen_final['total_pagado'].fillna(0)
        resumen_final['tarifa_efectiva'] = resumen_final.apply(
            lambda row: (row['total_pagado'] / row['total_horas']) if row['total_horas'] > 0 else 0, axis=1
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Operadores"
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Resumen Financiero de Operadores'
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:F2')
        fecha_ini_str = filtros.get('fecha_inicio', 'N/A').strftime('%Y-%m-%d') if filtros.get('fecha_inicio') else "Inicio"
        fecha_fin_str = filtros.get('fecha_fin', 'N/A').strftime('%Y-%m-%d') if filtros.get('fecha_fin') else "Fin"
        ws['A2'] = f"Período del reporte: {fecha_ini_str} al {fecha_fin_str}"
        ws['A2'].font = Font(size=11, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        df_export = resumen_final.copy()
        df_export.columns = ["Operador", "Equipo", "Total Horas", "Ingreso Generado", "Total Pagado", "Tarifa Efectiva"]
        for r in dataframe_to_rows(df_export, index=False, header=True): ws.append(r)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        formato_horas = '0.00'
        formato_moneda = f'"{moneda_symbol}" #,##0.00'
        for row_idx in range(4, ws.max_row + 1):
            ws[f'C{row_idx}'].number_format = formato_horas
            ws[f'D{row_idx}'].number_format = formato_moneda
            ws[f'E{row_idx}'].number_format = formato_moneda
            ws[f'F{row_idx}'].number_format = formato_moneda

        fila_inicio_totales = ws.max_row + 2
        ws.cell(row=fila_inicio_totales, column=2, value="TOTALES:").font = Font(bold=True)
        ws.cell(row=fila_inicio_totales, column=3, value=resumen_final['total_horas'].sum()).number_format = formato_horas
        ws.cell(row=fila_inicio_totales, column=4, value=resumen_final['total_ingresos'].sum()).number_format = formato_moneda
        ws.cell(row=fila_inicio_totales, column=5, value=resumen_final['total_pagado'].sum()).number_format = formato_moneda
        ws.cell(row=fila_inicio_totales, column=6, value="") # Tarifa efectiva total no aplica

        for col in ws.iter_cols(min_row=3):
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        try:
            wb.save(ruta_guardar)
            return True, ruta_guardar
        except PermissionError:
            return False, "No se pudo guardar el archivo. Asegúrate de que no esté abierto."
        except Exception as e:
            return False, str(e)
