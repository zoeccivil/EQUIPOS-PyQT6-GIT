import openpyxl

class XLSXUtils:
    """Utilidad para manejar archivos Excel (.xlsx)."""

    @staticmethod
    def leer_xlsx(nombre_archivo, hoja=None):
        """Lee un archivo XLSX y retorna los datos como lista de dicts."""
        wb = openpyxl.load_workbook(nombre_archivo, data_only=True)
        ws = wb[hoja] if hoja else wb.active
        encabezados = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            datos.append(dict(zip(encabezados, row)))
        return datos

    @staticmethod
    def escribir_xlsx(nombre_archivo, datos, encabezados, hoja="Hoja1"):
        """Escribe una lista de dicts en un archivo XLSX."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = hoja
        ws.append(encabezados)
        for d in datos:
            ws.append([d.get(c, "") for c in encabezados])
        wb.save(nombre_archivo)
        return nombre_archivo
